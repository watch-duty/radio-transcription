"""Geographic wildfire risk scoring.

Tries three data sources in order:
  1. FEMA NRI county CSV (auto-download)
  2. USDA Forest Service "Wildfire Risk to Communities" XLSX (auto-download)
  3. Hardcoded state heuristic (fallback)

Uses 5-digit county FIPS codes as keys, mapping to normalized 0-1 risk scores.
"""
import csv
import logging
import os
import sys

import requests

import cache
from config import (
    DEFAULT_RISK_SCORE,
    FEMA_NRI_URL,
    HIGH_RISK_SCORE,
    HIGH_RISK_STATES,
    STATE_FIPS_TO_ABBREV,
)

logger = logging.getLogger(__name__)

_CACHE_SOURCE = "fema"
_NRI_KEY = "nri_counties"
_WRC_KEY = "wrc_counties"

# USDA Forest Service Wildfire Risk to Communities (stable direct URL)
_WRC_URL = (
    "https://wildfirerisk.org/wp-content/uploads/2025/05/"
    "wrc_download_202505.xlsx"
)


def _download(url, cache_key, ext, expected_content_type=None):
    """Download a file, detect HTML responses, cache to disk."""
    cached_path = cache.cache_path(_CACHE_SOURCE, cache_key).with_suffix(ext)
    if cached_path.exists():
        return cached_path

    try:
        resp = requests.get(url, timeout=120, stream=True)
        resp.raise_for_status()

        cached_path.parent.mkdir(parents=True, exist_ok=True)
        with open(cached_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=65536):
                f.write(chunk)

        # Detect HTML bait-and-switch (FEMA URL serves a landing page)
        with open(cached_path, "rb") as f:
            head = f.read(64).lstrip().lower()
        if head.startswith(b"<!doctype") or head.startswith(b"<html"):
            logger.warning(
                "URL %s returned HTML (not %s) — skipping.", url, ext
            )
            cached_path.unlink()
            return None

        if expected_content_type:
            ct = resp.headers.get("Content-Type", "")
            if expected_content_type not in ct:
                logger.warning(
                    "URL %s returned %s, expected %s",
                    url, ct, expected_content_type,
                )
                cached_path.unlink()
                return None

        size_mb = os.path.getsize(cached_path) / (1024 * 1024)
        print(
            f"  geo_risk: downloaded {url.split('/')[-1]} ({size_mb:.1f} MB)",
            file=sys.stderr,
        )
        return cached_path
    except Exception as e:
        logger.warning("Download failed for %s: %s", url, e)
        if cached_path.exists():
            cached_path.unlink()
        return None


def _normalize(scores):
    """Min-max normalize a dict of scores to 0-1."""
    if not scores:
        return {}
    vals = list(scores.values())
    lo, hi = min(vals), max(vals)
    if hi == lo:
        return {k: 0.5 for k in scores}
    return {k: (v - lo) / (hi - lo) for k, v in scores.items()}


def _try_fema_nri():
    """Attempt to load FEMA NRI county data. Returns dict or None."""
    csv_path = _download(FEMA_NRI_URL, _NRI_KEY, ".csv")
    if csv_path is None:
        return None

    raw = {}
    try:
        with open(csv_path, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                fips = row.get("STCOFIPS", "").strip()
                val = row.get("WFIR_RISKS", "").strip()
                if fips and val:
                    try:
                        raw[fips] = float(val)
                    except ValueError:
                        pass
    except Exception as e:
        logger.warning("FEMA NRI parse failed: %s", e)
        return None

    if not raw:
        return None

    normalized = _normalize(raw)
    print(
        f"  geo_risk: loaded FEMA NRI data for {len(normalized)} counties",
        file=sys.stderr,
    )
    return normalized


def _try_wrc():
    """Attempt to load USDA WRC XLSX. Returns dict or None."""
    try:
        import openpyxl
    except ImportError:
        logger.warning("openpyxl not installed — skipping WRC")
        return None

    xlsx_path = _download(_WRC_URL, _WRC_KEY, ".xlsx")
    if xlsx_path is None:
        return None

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "Counties" not in wb.sheetnames:
            logger.warning("WRC XLSX missing 'Counties' sheet")
            return None

        ws = wb["Counties"]
        iter1 = ws.iter_rows(values_only=True)
        headers = next(iter1)
        # Find GEOID (5-digit FIPS) and RISK_NATIONAL_RANK columns
        try:
            geoid_idx = headers.index("GEOID")
            risk_idx = headers.index("RISK_NATIONAL_RANK")
        except ValueError:
            logger.warning(
                "WRC XLSX missing GEOID or RISK_NATIONAL_RANK column"
            )
            return None

        county_risk = {}
        for row in iter1:
            geoid = row[geoid_idx]
            risk = row[risk_idx]
            if geoid and risk is not None:
                fips = str(geoid).zfill(5)
                try:
                    county_risk[fips] = float(risk)
                except (ValueError, TypeError):
                    pass

        wb.close()
    except Exception as e:
        logger.warning("WRC parse failed: %s", e)
        return None

    if not county_risk:
        return None

    # RISK_NATIONAL_RANK is already 0-1, no normalization needed
    print(
        f"  geo_risk: loaded USDA WRC data for {len(county_risk)} counties",
        file=sys.stderr,
    )
    return county_risk


def _build_state_avg(county_risk):
    """Build per-state-abbrev average from county data."""
    sums = {}
    counts = {}
    for fips, score in county_risk.items():
        abbrev = STATE_FIPS_TO_ABBREV.get(fips[:2])
        if abbrev:
            sums[abbrev] = sums.get(abbrev, 0.0) + score
            counts[abbrev] = counts.get(abbrev, 0) + 1
    return {
        st: round(sums[st] / counts[st], 4)
        for st in sums if counts[st] > 0
    }


def load_nri_data():
    """Load county-level wildfire risk data, with multiple fallbacks.

    Returns:
        (county_risk, state_avg) where:
        - county_risk: dict[fips_str, float 0-1] or None
        - state_avg: dict[state_abbrev, float 0-1] or None
    """
    # Attempt 1: FEMA NRI (current URL often returns HTML; we detect it)
    county_risk = _try_fema_nri()

    # Attempt 2: USDA WRC
    if county_risk is None:
        county_risk = _try_wrc()

    if county_risk is None:
        print("  geo_risk: no county data — using state heuristic", file=sys.stderr)
        return None, None

    state_avg = _build_state_avg(county_risk)
    return county_risk, state_avg


def score_geo_risk(county_fips_str, state, county_risk, state_avg):
    """Score geographic wildfire risk for a feed."""
    # Try county-level lookup
    if county_risk and county_fips_str:
        fips_codes = [f.strip() for f in county_fips_str.split(",") if f.strip()]
        scores = [county_risk[f] for f in fips_codes if f in county_risk]
        if scores:
            return max(scores)

    # Try state-level average
    if state_avg and state:
        st = state.upper().strip()
        if st in state_avg:
            return state_avg[st]

    # Fallback: state heuristic
    if state:
        st = state.upper().strip()
        if st in HIGH_RISK_STATES:
            return HIGH_RISK_SCORE
        return DEFAULT_RISK_SCORE

    return 0.5
